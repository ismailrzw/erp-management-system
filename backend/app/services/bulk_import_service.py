"""Validation and atomic-style bulk creation of student accounts."""

import csv
from datetime import datetime, timezone
from io import TextIOWrapper
from uuid import uuid4

import bcrypt
from openpyxl import load_workbook
from werkzeug.datastructures import FileStorage

from app.extensions import mongo
from app.models.user import Role, UserFields
from app.schemas.student_schema import CreateStudentSchema
from app.services.student_service import (
    generate_initial_password,
    generate_student_email,
)

REQUIRED_COLUMNS = ("Name", "Roll", "Department", "Section", "Session", "Course", "Teacher", "Recovery Email")
_COLUMN_MAP = {column.casefold(): column for column in REQUIRED_COLUMNS}

HEADER_ALIASES = {
    "name": "Name",
    "student name": "Name",
    "full name": "Name",
    "roll": "Roll",
    "roll no": "Roll",
    "roll no.": "Roll",
    "roll number": "Roll",
    "roll_no": "Roll",
    "department": "Department",
    "dept": "Department",
    "section": "Section",
    "sec": "Section",
    "session": "Session",
    "academic session": "Session",
    "course": "Course",
    "course name": "Course",
    "teacher": "Teacher",
    "assigned teacher": "Teacher",
    "instructor": "Teacher",
    "recovery email": "Recovery Email",
    "recovery_email": "Recovery Email",
    "personal email": "Recovery Email",
    "email": "Recovery Email",
}


def _normalise_row(row: dict) -> dict:
    return {
        "name": str(row.get("Name") or "").strip(),
        "roll": str(row.get("Roll") or "").strip(),
        "dept": str(row.get("Department") or "").strip(),
        "section": str(row.get("Section") or "").strip(),
        "session": str(row.get("Session") or "").strip(),
        "course": str(row.get("Course") or "").strip(),
        "teacher": str(row.get("Teacher") or "").strip(),
        "recovery_email": str(row.get("Recovery Email") or "").strip() or None,
    }


def _read_rows(file: FileStorage) -> list[dict]:
    filename = (file.filename or "").lower()
    if filename.endswith(".csv"):
        wrapper = TextIOWrapper(file.stream, encoding="utf-8-sig", newline="")
        return list(csv.DictReader(wrapper))
    if filename.endswith(".xlsx"):
        workbook = load_workbook(file.stream, read_only=True, data_only=True)
        sheet = workbook.active
        values = sheet.iter_rows(values_only=True)
        headers = next(values, None)
        if not headers:
            return []
        headers = [str(value).strip() if value is not None else "" for value in headers]
        return [dict(zip(headers, row)) for row in values if any(value is not None and str(value).strip() for value in row)]
    raise ValueError("Only .csv and .xlsx files are supported.")


def parse_excel(file: FileStorage) -> tuple[list[dict], list[dict]]:
    """Read CSV/XLSX, require all columns, and validate every non-empty row."""
    if not file or not file.filename:
        raise ValueError("An import file is required.")
    rows = _read_rows(file)
    if not rows:
        raise ValueError("The import file has no data rows.")
    
    # Map raw headers to canonical column names using HEADER_ALIASES
    raw_header_map = {}
    for key in rows[0].keys():
        if key is None:
            continue
        cleaned = str(key).strip().casefold()
        canonical = HEADER_ALIASES.get(cleaned, cleaned)
        raw_header_map[canonical.casefold()] = key

    missing = [column for column in REQUIRED_COLUMNS if column.casefold() not in raw_header_map]
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(missing)}.")

    schema = CreateStudentSchema()
    valid_rows, errors, seen_rolls = [], [], set()
    for index, raw_row in enumerate(rows, start=2):
        canonical = {required: raw_row.get(raw_header_map[required.casefold()]) for required in REQUIRED_COLUMNS}
        data = _normalise_row(canonical)
        try:
            data = schema.load(data)
        except Exception as exc:  # noqa: BLE001 - must not let one bad row abort the whole import
            errors.append({"row": index, "error": str(exc)})
            continue
        roll_key = data["roll"].casefold()
        if roll_key in seen_rolls:
            errors.append({"row": index, "roll": data["roll"], "error": "Duplicate roll in import file."})
            continue
        seen_rolls.add(roll_key)
        valid_rows.append({"row": index, "data": data})
    return valid_rows, errors


def bulk_create_students(rows: list[dict]) -> dict:
    """Create valid students in one batch; clean up the batch if insertion fails."""
    if not rows:
        return {"imported_count": 0, "skipped_count": 0, "errors": []}

    rolls = [row["data"]["roll"] for row in rows]
    existing_rolls = {
        value.casefold()
        for value in mongo.db.users.distinct(UserFields.ROLL, {UserFields.ROLL: {"$in": rolls}})
        if isinstance(value, str)
    }
    batch_id = uuid4().hex
    documents, errors = [], []
    now = datetime.now(timezone.utc)
    for row in rows:
        data = row["data"]
        if data["roll"].casefold() in existing_rolls:
            errors.append({"row": row["row"], "roll": data["roll"], "error": "Roll already exists."})
            continue
        password = generate_initial_password(data["roll"])
        documents.append({
            UserFields.NAME: data["name"],
            UserFields.EMAIL: generate_student_email(data["roll"]),
            UserFields.PASSWORD_HASH: bcrypt.hashpw(password.encode("utf-8"), bcrypt.gensalt()).decode("utf-8"),
            UserFields.ROLE: Role.STUDENT,
            UserFields.DEPT: data["dept"].upper(),
            UserFields.SECTION: data["section"].upper(),
            UserFields.COURSE: data["course"],
            UserFields.ROLL: data["roll"],
            "session": data["session"],
            "teacher": data["teacher"],
            UserFields.RECOVERY_EMAIL: data["recovery_email"],
            UserFields.DELETED: False,
            UserFields.DELETED_AT: None,
            UserFields.CREATED_AT: now,
            UserFields.UPDATED_AT: now,
            "import_batch_id": batch_id,
        })
    if not documents:
        return {"imported_count": 0, "skipped_count": len(errors), "errors": errors}
    try:
        mongo.db.users.insert_many(documents, ordered=True)
    except Exception as exc:
        mongo.db.users.delete_many({"import_batch_id": batch_id})
        raise ValueError(f"Import failed and was rolled back: {exc}") from exc
    mongo.db.users.update_many({"import_batch_id": batch_id}, {"$unset": {"import_batch_id": ""}})
    return {"imported_count": len(documents), "skipped_count": len(errors), "errors": errors}
